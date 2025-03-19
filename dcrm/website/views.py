import openpyxl
import re
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .forms import AddRecordForm
from .models import Record
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook

def home(request):
    # Pobieranie wszystkich rekordów
    records = Record.objects.all()

    # Sprawdzanie, czy wykonano wyszukiwanie
    query = request.GET.get('q', '')  # Pobiera wartość wyszukiwania z parametru GET
    if query:
        # Filtrowanie rekordów na podstawie 'id' lub 'client_name'
        records = records.filter(id__icontains=query) | records.filter(client_name__icontains=query)

    # Sprawdzanie logowania
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        # Autentykacja
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Zalogowano!")
            return redirect('home')
        else:
            messages.success(request, "Wystąpił błąd podczas logowania. Spróbuj ponownie...")
            return redirect('home')
    else:
        # Renderowanie widoku z rekordami
        return render(request, 'home.html', {'records': records, 'query': query})

def logout_user(request):
    logout(request)
    messages.success(request, "Zostałeś wylogowany...")
    return redirect('home')

def work_in_progress(request):
        return render(request, 'work_in_progress.html', {})

def customer_record(request, pk):
    if request.user.is_authenticated:
        customer_record = Record.objects.get(id=pk)
        return render(request, 'record.html', {'customer_record':customer_record})
    else:
        messages.success(request, "Musisz się zalogować ...")
        return redirect('home')

def delete_record(request, pk):
    if request.user.is_authenticated:
        delete_id = Record.objects.get(id=pk)
        delete_id.delete()
        messages.success(request, "Wycena usunięta poprawnie...")
        return redirect('home')
    else:
        messages.success(request, "Musisz się zalogować...")
        return render(request, 'home.html')

def add_record(request):
    form = AddRecordForm(request.POST or None)
    if request.user.is_authenticated:
        if request.method == "POST":
            vcpu = request.POST.get("vcpu")
            vram = request.POST.get("vram")
            disk = request.POST.get("disk")
            pbs = request.POST.get("pbs")
            pbs_replication = request.POST.get("pbs_replication")

            # Sprawdzanie czy wartości są liczbami przed konwersją do int
            if vcpu and not vcpu.isdigit():
                messages.error(request, "vCPU musi być liczbą!")
                return render(request, 'add_record.html', {'form': form})

            if vram and not vram.isdigit():
                messages.error(request, "vRAM musi być liczbą!")
                return render(request, 'add_record.html', {'form': form})

            if disk and not disk.isdigit():
                messages.error(request, "Dysk musi być liczbą!")
                return render(request, 'add_record.html', {'form': form})

            if pbs and not pbs.isdigit():
                messages.error(request, "PBS musi być liczbą!")
                return render(request, 'add_record.html', {'form': form})
            
            if pbs_replication and not pbs_replication.isdigit():
                messages.error(request, "Replikacja PBS musi być liczbą!")
                return render(request, 'add_record.html', {'form': form})

            vcpu = int(vcpu) if vcpu else 0
            vram = int(vram) if vram else 0
            disk = int(disk) if disk else 0
            pbs = int(pbs) if pbs else 0
            pbs_replication = int(pbs_replication) if pbs_replication else 0

            # Sprawdzenie, czy vCPU i vRAM są zgodne z wymaganiami
            if vcpu > 32:
                messages.error(request, "Maksymalna dozwolona wartość vCPU to 32!")
                return render(request, 'add_record.html', {'form': form})

            if vram > 32:
                messages.error(request, "Maksymalna dozwolona wartość vRAM to 32!")
                return render(request, 'add_record.html', {'form': form})

            if disk < 20 or disk > 1000:
                messages.error(request, "Rozmiar dysku możliwy do wyboru to zakres od 20 do 1000!")
                return render(request, 'add_record.html', {'form': form})

            # Sprawdzenie, czy PBS jest taki sam jak Disk
            if pbs != disk:
                messages.error(request, "PBS musi być równy wartości dysku!")
                return render(request, 'add_record.html', {'form': form})

            if form.is_valid():
                form.save()
                messages.success(request, "Wycena dodana...")
                return redirect('home')

            if pbs_replication not in [0, pbs]:
                messages.error(request, "Replikacja PBS może wynosić 0 lub być równa wartości PBS!")
                return render(request, 'add_record.html', {'form': form})

        return render(request, 'add_record.html', {'form': form})
    else:
        messages.error(request, "Musisz się zalogować...")
        return redirect('home')

def update_record(request, pk):
    if request.user.is_authenticated:
        current_record = Record.objects.get(id=pk)
        previous_is_accepted = current_record.is_accepted
        
        form = AddRecordForm(request.POST or None, instance=current_record)
        
        if request.method == "POST":
            vcpu = request.POST.get("vcpu")
            vram = request.POST.get("vram")
            disk = request.POST.get("disk")
            pbs = request.POST.get("pbs")
            pbs_replication = request.POST.get("pbs_replication")

            if vcpu and not vcpu.isdigit():
                messages.error(request, "vCPU musi być liczbą!")
                return render(request, 'update_record.html', {'form': form})

            if vram and not vram.isdigit():
                messages.error(request, "vRAM musi być liczbą!")
                return render(request, 'update_record.html', {'form': form})

            if disk and not disk.isdigit():
                messages.error(request, "Dysk musi być liczbą!")
                return render(request, 'update_record.html', {'form': form})

            if pbs and not pbs.isdigit():
                messages.error(request, "PBS musi być liczbą!")
                return render(request, 'update_record.html', {'form': form})

            if pbs_replication and not pbs_replication.isdigit():
                messages.error(request, "Replikacja PBS musi być liczbą!")
                return render(request, 'update_record.html', {'form': form})

            vcpu = int(vcpu) if vcpu else 0
            vram = int(vram) if vram else 0
            disk = int(disk) if disk else 0
            pbs = int(pbs) if pbs else 0
            pbs_replication = int(pbs_replication) if pbs_replication else 0

            if vcpu > 32:
                messages.error(request, "Maksymalna dozwolona wartość vCPU to 32!")
                return render(request, 'update_record.html', {'form': form})

            if vram > 32:
                messages.error(request, "Maksymalna dozwolona wartość vRAM to 32!")
                return render(request, 'update_record.html', {'form': form})

            if disk < 20 or disk > 1000:
                messages.error(request, "Rozmiar dysku możliwy do wyboru to zakres od 20 do 1000!")
                return render(request, 'update_record.html', {'form': form})

            if pbs != disk:
                messages.error(request, "PBS musi być równy wartości dysku!")
                return render(request, 'update_record.html', {'form': form})

            if pbs_replication not in [0, pbs]:
                messages.error(request, "Replikacja PBS może wynosić 0 lub być równa wartości PBS!")
                return render(request, 'update_record.html', {'form': form})

            if form.is_valid():
                updated_record = form.save(commit=False)
                if previous_is_accepted == updated_record.is_accepted:
                    updated_record.is_accepted = 0
                updated_record.save()
                messages.success(request, "Wycena zmodyfikowana...")
                return redirect('home')
        
        return render(request, 'update_record.html', {'form': form})
    else:
        messages.error(request, "Musisz się zalogować...")
        return redirect('home')

#Testy generowania wyceny

def generate_xlsx(request, record_id):
    # Pobranie rekordu z bazy danych
    customer_record = Record.objects.get(id=record_id)

    # Ścieżka do szablonu
    template_path = "website/templates/template_vps.xlsx"

    # Otwieramy plik Excel jako szablon
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Dodajemy nazwę klienta do A1
    ws["A1"] = customer_record.client_name

    # Sprawdzamy wartość procedury i ustawiamy odpowiednią wartość w B1
    if customer_record.procedure == "Nowy VPS":
        ws["B1"] = "Nowy VPS"
    elif customer_record.procedure == "Rozbudowa obecnego VPS":
        ws["B1"] = "Rozbudowa obecnego VPS"

    # Mapa wartości do podmiany (upewniamy się, że wartości są liczbami)
    data_map = {
        "vCPU": int(customer_record.vcpu) if customer_record.vcpu is not None else 0,
        "RAM": int(customer_record.vram) if customer_record.vram is not None else 0,
        "Dysk": int(customer_record.disk) if customer_record.disk is not None else 0,
        "Backup": int(customer_record.pbs) if customer_record.pbs is not None else 0,
        "Strefa bezpieczeństwa": int(customer_record.dmz) if customer_record.dmz is not None else 0,
        "Sieć wewnętrzna - 1Gbit/s": int(customer_record.network) if customer_record.network is not None else 0,
        "Replikacja Backupu": int(customer_record.pbs_replication) if customer_record.pbs_replication is not None else 0,
        "vConnect": int(customer_record.vconnect) if customer_record.vconnect is not None else 0,
        "Adresacja IPv4": customer_record.ip,
        "Godziny administracyjne (h) - Tier III": float(customer_record.adm_hours) if customer_record.adm_hours is not None else 0.0,
        "Windows Server 2022 Standard - 8 Core License Pack 1 Year": int(customer_record.ws2022_1) if customer_record.ws2022_1 is not None else 0,
        "Windows Server 2022 Standard - 8 Core License Pack 3 Year": int(customer_record.ws2022_3) if customer_record.ws2022_3 is not None else 0,
        "Windows Server 2022 CAL - 1 User CAL - 1 year": int(customer_record.ws2022_cal_1) if customer_record.ws2022_cal_1 is not None else 0,
        "Windows Server 2022 CAL - 1 User CAL - 3 year": int(customer_record.ws2022_cal_3) if customer_record.ws2022_cal_3 is not None else 0,
        "Windows Server 2022 Remote Desktop Services - 1 User CAL 1 Year": int(customer_record.rds_cal_1) if customer_record.rds_cal_1 is not None else 0,
        "Windows Server 2022 Remote Desktop Services - 1 User CAL 3 Year": int(customer_record.rds_cal_3) if customer_record.rds_cal_3 is not None else 0,
        "Windows Server 2022 Remote Desktop Services - 1 Device CAL - perpetual": int(customer_record.rds_cal_perpetual) if customer_record.rds_cal_perpetual is not None else 0,
        "Data Space Shield - Firewall Premium - 10 reguł": int(customer_record.fw_premium) if customer_record.fw_premium is not None else 0,
        "Data Space Shield - Geo Firewall - 1 kraj": int(customer_record.geofw) if customer_record.geofw is not None else 0,
        "Data Space Shield - IPSEC": int(customer_record.ipsec) if customer_record.ipsec is not None else 0,
        "Data Space Shield - SSL VPNs": int(customer_record.ssl_vpn) if customer_record.ssl_vpn is not None else 0,
        "Data Space Shield - Guard DNS": int(customer_record.dns_guard) if customer_record.dns_guard is not None else 0,
        "Data Space Shield - Webfiltering": int(customer_record.webfiltering) if customer_record.webfiltering is not None else 0,
        "Data Space Shield - IDS/IPS/AV - 100Mbit/s": int(customer_record.ids_ips) if customer_record.ids_ips is not None else 0,
        "Data Space Shield - vDOM": int(customer_record.vdom) if customer_record.vdom is not None else 0,
    }

    # Podmiana wartości w kolumnie C na podstawie nazw w kolumnie A
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell_a = row[0]  # Komórka w kolumnie A
        if cell_a.value in data_map:
            cell_c = row[2]  # Komórka w kolumnie C (indeks 2)
            cell_c.value = data_map[cell_a.value]  # Wstawienie wartości

    # Obsługa profilu wydajnościowego dysku
    disk_profile_row = {1: 10, 2: 11, 3: 12}  # Mapowanie profilu na wiersz w Excelu
    if customer_record.disk_profile in disk_profile_row:
        row_idx = disk_profile_row[customer_record.disk_profile]
        ws[f"C{row_idx}"] = 1  # Wstawienie wartości 1 do odpowiedniego wiersza

    # Obsługa zmiennej dmz
    if customer_record.dmz == '0':
        ws["C14"].value = 1
        ws["C15"].value = None  # Usunięcie wartości z C15, jeśli wcześniej coś tam było
    elif customer_record.dmz == '1':
        ws["C15"].value = 1
        ws["C14"].value = None  # Usunięcie wartości z C14, jeśli wcześniej coś tam było

    # Przygotowanie pliku do pobrania
    client_name = re.sub(r'[^a-zA-Z0-9_-]', '_', customer_record.client_name)  # Usunięcie znaków specjalnych
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=wycena_vps_{client_name}.xlsx'

    wb.save(response)
    return response